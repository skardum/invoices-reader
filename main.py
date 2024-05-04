import sys
import os
import cv2
import pyzbar.pyzbar as pyzbar
import base64
import openpyxl
import re
from dotenv import load_dotenv
from PyQt6.QtWidgets import QApplication, QMainWindow, QPushButton, QLabel, QVBoxLayout, QGraphicsView, QWidget, QDialog, QMessageBox, QFileDialog, QGraphicsScene, QGraphicsPixmapItem
from PyQt6 import QtCore
from PyQt6.QtCore import pyqtSignal, Qt, QObject, QThread
from PyQt6.QtGui import QPixmap, QImage
from PyQt6.uic import loadUi
from threading import Thread
import PIL.Image
from datetime import datetime
from database import connect_to_database, save_detection_to_database, DatabaseDialog, insert_extracted_data
import google.generativeai as genai
import time  # Import time module for generating unique detection IDs
import sqlite3


# Load environment variables
load_dotenv()

# Configure Google API
os.getenv("GOOGLE_API_KEY")
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))


def process_text_and_fill_ui(image_data):
    try:
        model = genai.GenerativeModel('gemini-pro-vision')
        response = model.generate_content(['''Extract the following fields from this invoice and format the information as a dictionary:
                                           - Vendor Name
                                           - Vendor VAT ID
                                           - Date
                                           - Total Amount
                                           - VAT Amount
                                           - invoice number

                                           Please provide the extracted information in the format:
                                           {
                                               'vendor_name': '...',
                                               'vat_id': '...',
                                               'date': '...',
                                               'invoice_total': '...',
                                               'vat_total': '...',
                                               'invoice_number': '...'
                                           }''',
                                           image_data
                                           ])

        print(response.text)
        pattern = r'{.*?}'
        match = re.search(pattern, response.text, re.DOTALL)
        if match:
            print(match)
            dictionary_text = match.group()
            parsed_data = eval(dictionary_text)
            return parsed_data
        else:
            print("No dictionary found in the gemini output.")
            return None
    except Exception as e:
        print(f"Error during gemini processing: {e}")
        return None


def update_ui_with_data(ui, data):
    ui.vendor_lineedit.setText(data.get('vendor_name', ''))
    ui.vatid_lineedit.setText(data.get('vat_id', ''))
    ui.date_lineedit.setText(data.get('date', ''))

    # Convert float to string before setting the text
    invoice_total = str(data.get('invoice_total', ''))
    ui.total_lineedit.setText(invoice_total)

    # Convert float to string before setting the text
    vat_total = str(data.get('vat_total', ''))
    ui.vatamount_lineedit.setText(vat_total)
    invoice_number = str(data.get('invoice_number', ''))
    ui.invoicenumber_lineedit.setText(invoice_number)


dark_theme_stylesheet = """
QWidget {
    background-color: #333;
    color: #fff;
}

QPushButton {
    background-color: #555;
    color: #fff;
    border: 2px solid #444;
    border-radius: 5px;
    padding: 5px 10px;
}

QPushButton:hover {
    background-color: #666;
}

QLineEdit, QTextEdit {
    background-color: #444;
    color: #fff;
    border: 1px solid #222;
    border-radius: 5px;
    padding: 5px;
}

QLabel {
    color: #fff;
}

QProgressBar {
    background-color: #444;
    color: #fff;
    border: 1px solid #222;
    border-radius: 5px;
}

QProgressBar::chunk {
    background-color: #666;
}
"""


class ProgressDialog(QDialog):
    def __init__(self, parent=None):
        super(ProgressDialog, self).__init__(parent)
        loadUi('progress_dialog.ui', self)  # Load the UI file
        self.setModal(True)
        self.setWindowTitle("Processing")
        self.progressBar.setRange(0, 0)  # Indeterminate progress bar


class AiExtractor(QObject):
    progress_changed = pyqtSignal(int, str)  # Signal for progress updates
    started = pyqtSignal()  # Signal for indicating start of processing
    # Signal to emit extracted data and image path
    data_extracted = pyqtSignal(dict, str)

    def __init__(self, image_path, ui, db_connection, parent=None):
        super(AiExtractor, self).__init__(parent)
        self.image_path = image_path  # Set image_path
        self.ui = ui
        self.db_connection = db_connection

    def extract(self):
        self.started.emit()  # Emit signal indicating start of processing

        try:
            # Read the image file
            with open(self.image_path, 'rb') as file:
                image_data = PIL.Image.open(file)
                self.progress_changed.emit(
                    25, "Image loaded")  # Emit progress update
                parsed_data = process_text_and_fill_ui(image_data)
                if parsed_data:
                    self.progress_changed.emit(
                        50, "Data extracted")  # Emit progress update
                    update_ui_with_data(self.ui, parsed_data)
                    # Save data to database
                    # Emit signal with extracted data and image path
                    self.data_extracted.emit(parsed_data, self.image_path)
                    self.progress_changed.emit(
                        100, "Extraction complete")  # Emit progress update
                else:
                    self.progress_changed.emit(
                        0, "No data extracted")  # Emit progress update
        except FileNotFoundError:
            self.progress_changed.emit(
                0, "File not found")  # Emit progress update
        except Exception as e:
            self.progress_changed.emit(
                0, f"Error: {str(e)}")  # Emit progress update


def remove_non_printable(text):
    """
    تنظف النص المعطى بإزالة الأحرف غير القابلة للطباعة وبعض الأحرف الخاصة،
    وتستبدلها بفواصل، ثم تزيل الفواصل المكررة وتقسم النص إلى قائمة بناءً على الفواصل.

    Args:
    text (str): النص الذي يحتاج إلى تنظيف.

    Returns:
    dict: قاموس يحتوي على المعلومات المطلوبة.
    """
    # استبدال الأحرف الخاصة بفواصل
    chars_to_remove = "?,☺,☻,♥,♦,♠,♣,§,¶,+,="
    for char in chars_to_remove:
        text = text.replace(char, ",")

    # إزالة الفاصلة من بداية النص إذا كانت موجودة
    if len(text) > 0 and text[0] == ",":
        text = text[1:]

    # إزالة الفواصل المكررة
    text = ','.join(text.split(","))
    cleaned_string = text.replace("\x01", ",").replace("\x02", ",").replace(
        "\x03", ",").replace("\x04", ",").replace("\x05", ",").replace("\x0e", ",").replace("\x1e", ",").replace("\x13", ",").replace("\x1f", ",").replace("\x1a", ",").replace("\x11", ",").replace("\x10", ",").replace("\x1c", ",")
    cleaned_string = cleaned_string.replace("\x0f", ",").replace(
        "\x14", ",").replace("\x15", ",").replace("\x06", ",").replace("\x07", ",").replace("\x19", ",").replace("\x08", ",").replace("\t", ",").replace(".00", "").replace("\x16", ",").replace("\x0b", ",").replace("\x0c", ",").replace("\x1b", ",")

    # تقسيم النص المنظف إلى قائمة بناءً على الفواصل
    result = cleaned_string.split(",")
    result2 = list(filter(None, result))
    # إنشاء قاموس من القائمة بفرض ترتيب ثابت
    print(result2)
    return {
        'vendor_name': result2[0] if len(result2) > 0 else '',
        'date': result2[1] if len(result2) > 1 else '',
        'vat_id': result2[2] if len(result2) > 2 else '',
        'invoice_total': result2[3] if len(result2) > 3 else '',
        'vat_total': result2[4] if len(result2) > 4 else ''
    }


def decode_qr_code(image):
    # Convert image to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Define a list of parameter combinations to try for adaptive thresholding
    adaptive_threshold_params = [
        # Combination 1
        (cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 15, 2),
        # Combination 2
        (cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 15, 2),
        # Combination 3
        (cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 25, 3),
        # Combination 4
        (cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 25, 3),
        # Combination 5
        (cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 35, 4),
        # Combination 6
        (cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 35, 4),
        # Add more parameter combinations as needed
    ]

    # Define a list of additional preprocessing settings to try
    preprocessing_settings = [
        # Preprocessing 1: No additional preprocessing
        lambda img: img,
        # Preprocessing 2: Gaussian blur
        lambda img: cv2.GaussianBlur(img, (5, 5), 0),
        # Preprocessing 3: Median blur
        lambda img: cv2.medianBlur(img, 5),
        # Preprocessing 4: Bilateral filter
        lambda img: cv2.bilateralFilter(img, 9, 75, 75),
        # Preprocessing 5: Adaptive histogram equalization
        lambda img: cv2.equalizeHist(img),
        # Add more preprocessing settings as needed
    ]

    for preprocess_func in preprocessing_settings:
        try:
            # Apply additional preprocessing
            processed_img = preprocess_func(gray)

            for params in adaptive_threshold_params:
                try:
                    # Apply adaptive thresholding with current parameters
                    threshold = cv2.adaptiveThreshold(
                        processed_img, 255, *params)

                    # Decode QR codes using PyZbar
                    qr_detections = pyzbar.decode(threshold)

                    for detection in qr_detections:
                        data = detection.data
                        try:
                            # Attempt to decode QR code data
                            text = base64.b64decode(data).decode("utf-8")
                            return text, threshold
                        except UnicodeDecodeError:
                            try:
                                # محاولة فك التشفير باستخدام ISO-8859-1
                                text = base64.b64decode(data).decode('iso-8859-1')
                            except UnicodeDecodeError:
                                try:
                                    # محاولة فك التشفير باستخدام windows-1256
                                    text = base64.b64decode(data).decode('windows-1256')

                                except Exception as e:
                                    # Handle decoding errors
                                    print(f"Error decoding QR code data: {e}")
                                    continue
                except Exception as e:
                    # Handle thresholding errors
                    print(f"Error applying adaptive thresholding: {e}")
                    continue
        except Exception as e:
            # Handle preprocessing errors
            print(f"Error applying preprocessing: {e}")
            continue

    # Return None if QR code is not detected
    return None, None


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        loadUi('qrdetector.ui', self)  # Load the UI file
        self.folder_path = None
        self.location = None
        self.sheet = None
        self.browseButton1.clicked.connect(self.choose_folder)
        self.browseButton3.clicked.connect(self.save_location)
        self.convertButton.clicked.connect(self.convert)
        self.save.clicked.connect(self.save_invoice)
        self.last.clicked.connect(self.load_last_invoice)
        self.next.clicked.connect(self.load_next_invoice)
        self.previous.clicked.connect(self.load_previous_invoice)
        self.first.clicked.connect(self.load_first_invoice)
        self.previous_extractions.clicked.connect(self.show_database_form)
        self.current_row = 2  # Start from row 2 to skip header
        self.setStyleSheet(dark_theme_stylesheet)
        self.current_detection_id = self.generate_unique_detection_id()

        # Connect the AI extract button to the new model
        self.ai_button.clicked.connect(self.ai_extract)

        # Initialize progress dialog
        self.progress_dialog = None

        # Connect to the database
        self.db_connection = connect_to_database()
        self.db_connection.row_factory = sqlite3.Row  # Set row factory to Row
        self.cursor = self.db_connection.cursor()
        self.image_path = None
        self.new_batch.clicked.connect(self.create_new_batch)

        # Initialize records list and current record index for navigation
        self.records = []
        self.current_record_index = 0

    def show_database_form(self):
        database_dialog = DatabaseDialog(self)
        database_dialog.exec()

    def create_new_batch(self):
        # Clear all inputs
        self.vendor_lineedit.clear()
        self.vatid_lineedit.clear()
        self.date_lineedit.clear()
        self.total_lineedit.clear()
        self.vatamount_lineedit.clear()
        self.invoicenumber_lineedit.clear()

        # Create a new unique detection ID
        self.current_detection_id = self.generate_unique_detection_id()

    def generate_unique_detection_id(self):
        # Generate a unique ID (example method, modify as needed)
        detection_id = int(time.time())  # Using current time as a simple unique ID
        print(f"New detection ID: {detection_id}")
        return detection_id

    def ai_extract(self):
        # Get the currently displayed pixmap from the graphics view
        pixmap = self.graphicsView.scene().items()[0].pixmap()

        # Convert pixmap to image
        image = pixmap.toImage()

        # Save the image to a temporary file
        image_path = "temp_image.jpg"
        image.save(image_path)

        # Create an instance of AiExtractor and move it to a separate thread
        self.ai_extractor = AiExtractor(image_path, self, self.db_connection)
        self.ai_thread = QtCore.QThread()
        self.ai_extractor.moveToThread(self.ai_thread)

        # Connect signals
        self.ai_extractor.progress_changed.connect(self.update_progress)
        self.ai_extractor.started.connect(self.show_progress_dialog)
        self.ai_thread.started.connect(self.ai_extractor.extract)

        # Start the thread
        self.ai_thread.start()

    def show_progress_dialog(self):
        # Show progress dialog when processing starts
        self.progress_dialog = ProgressDialog()
        self.progress_dialog.show()

    def update_progress(self, value, message):
        # Update progress dialog and UI with progress information
        self.progress_dialog.progressBar.setValue(value)
        self.progress_dialog.progressLabel.setText(message)

        # Close progress dialog when progress reaches 100
        if value == 100:
            self.progress_dialog.close()

    def load_invoice_data(self, row):
        try:
            # Assuming 'row' corresponds to 'id' in your database
            self.cursor.execute("SELECT * FROM extracted_data WHERE id=?", (row,))
            record = self.cursor.fetchone()
            if record:
                self.display_record(record)
            else:
                QMessageBox.warning(self, 'Error', 'No record found!')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to load data: {str(e)}')

    def save_invoice(self):
        try:
            current_record = self.records[self.current_record_index]
            self.cursor.execute("""
                UPDATE extracted_data
                SET vendor_name=?, vat_id=?, date=?, invoice_total=?, vat_total=?, invoice_number=?
                WHERE id=?
            """, (
                self.vendor_lineedit.text(),
                self.vatid_lineedit.text(),
                self.date_lineedit.text(),
                float(self.total_lineedit.text()),
                float(self.vatamount_lineedit.text()),
                self.invoicenumber_lineedit.text(),
                current_record['id']
            ))
            self.db_connection.commit()
            QMessageBox.information(self, 'Information', 'Invoice updated successfully!')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Failed to update invoice: {str(e)}')

    def display_record(self, record):
        # Update UI elements with data from the record
        self.vendor_lineedit.setText(record['vendor_name'])
        self.vatid_lineedit.setText(record['vat_id'])
        self.date_lineedit.setText(record['date'])
        self.total_lineedit.setText(str(record['invoice_total']))
        self.vatamount_lineedit.setText(str(record['vat_total']))
        self.invoicenumber_lineedit.setText(record['invoice_number'])

        # Ensure QGraphicsScene is initialized
        if not self.graphicsView.scene():
            self.graphicsView.setScene(QGraphicsScene())

        # Now proceed to clear and update the scene
        self.graphicsView.scene().clear()
        image_path = record['image_file']
        if os.path.exists(image_path):
            pixmap = QPixmap(image_path)
            self.graphicsView.scene().addPixmap(pixmap)
            self.graphicsView.fitInView(self.graphicsView.scene().itemsBoundingRect(),
                                        Qt.AspectRatioMode.KeepAspectRatio)
        else:
            QMessageBox.warning(self, 'Error', 'Image file not found at the specified path.')

    def load_next_invoice(self):
        if self.current_record_index < len(self.records) - 1:
            self.current_record_index += 1
            self.display_record(self.records[self.current_record_index])

    def load_previous_invoice(self):
        if self.current_record_index > 0:
            self.current_record_index -= 1
            self.display_record(self.records[self.current_record_index])

    def load_first_invoice(self):
        if self.records:
            self.current_record_index = 0
            self.display_record(self.records[self.current_record_index])

    def load_last_invoice(self):
        if self.records:
            self.current_record_index = len(self.records) - 1
            self.display_record(self.records[self.current_record_index])

    def choose_folder(self):
        folder_path = QFileDialog.getExistingDirectory(
            self, "Select invoices Images Folder")
        if folder_path:
            self.folder_path = folder_path
            self.folder_path_label.setText(folder_path)

    def save_location(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Select excel file Location", "", "Excel Files (*.xlsx);;All Files (*)")
        if file_path:
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            self.location = file_path
            self.location_label.setText(file_path)

    def convert(self):
        folder_path = self.folder_path
        location = self.location

        if not folder_path or not location:
            QMessageBox.warning(
                self, 'Error', 'You need to select a folder and a save location!')
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "QR code data"
            sheet.append(["Image File", "Name of Seller",
                          "VAT Number", "Date and Time", "Total Amount", "VAT Amount", "Invoice Number"])

            num_files = len(os.listdir(folder_path))
            self.progressBar.setMaximum(num_files)

            # إنشاء سجل في جدول main_records واستخدام self.current_detection_id
            save_detection_to_database(
                self.db_connection, self.current_detection_id, num_files)

            if self.current_detection_id is None:
                QMessageBox.warning(
                    self, 'Warning', 'Failed to create a detection record in the database.')
                return
            processed_files = 0
            for file_name in os.listdir(folder_path):
                file_path = os.path.join(folder_path, file_name)
                if os.path.isfile(file_path):
                    image = cv2.imread(file_path)
                    invoice_data, threshold = decode_qr_code(image)
                    if invoice_data:
                        invoice_data_dict = remove_non_printable(invoice_data)
                        invoice_data_dict['image_path'] = file_path
                        invoice_data_dict['invoice_number'] = "0"  # تحديث هذا حسب الحاجة

                        # استخدام نفس detection_id لكل فاتورة
                        insert_extracted_data(
                            self.db_connection, self.current_detection_id, invoice_data_dict['image_path'], invoice_data_dict['vendor_name'], invoice_data_dict['date'], invoice_data_dict['vat_id'], invoice_data_dict['invoice_total'], invoice_data_dict['vat_total'], invoice_data_dict['invoice_number'])
                        sheet.append(
                            [invoice_data_dict['image_path'], invoice_data_dict['vendor_name'], invoice_data_dict['vat_id'], invoice_data_dict['date'], invoice_data_dict['invoice_total'], invoice_data_dict['vat_total'], invoice_data_dict['invoice_number']])
                    else:
                        sheet.append(
                            [file_path, "qr not detected", 0, 0, 0, 0, 0])
                        insert_extracted_data(
                            self.db_connection, self.current_detection_id, file_path, "qr not detected", 0, 0, 0, 0, 0)
                processed_files += 1
                progress_percent = int((processed_files / num_files) * 100)
                self.progressBar.setValue(progress_percent)  # Update the progress bar

            workbook.save(location)
            QMessageBox.information(
                self, 'Information', 'All invoices processed successfully!')

            # Automatically load the first invoice after saving
            self.current_row = 2
            self.load_invoice_data(self.current_row)  # Load data after saving
            print("Data loaded successfully!")
        except Exception as e:
            QMessageBox.critical(
                self, 'Error', f'Invoices read failed: {str(e)}')

    def save_invoice(self):
        try:
            if self.location is None or not os.path.exists(self.location):
                QMessageBox.warning(
                    self, 'Error', 'Please select a valid Excel file location!')
                return

            workbook = openpyxl.load_workbook(self.location)
            sheet = workbook["QR code data"]

            row = self.current_row

            vendor_name = self.vendor_lineedit.text()
            vat_id = self.vatid_lineedit.text()
            date = self.date_lineedit.text()
            invoice_total = self.total_lineedit.text()
            vat_total = self.vatamount_lineedit.text()
            invoice_number = self.invoicenumber_lineedit.text()

            sheet.cell(row=row, column=2, value=vendor_name)
            sheet.cell(row=row, column=3, value=vat_id)
            sheet.cell(row=row, column=4, value=date)
            sheet.cell(row=row, column=5, value=invoice_total)
            sheet.cell(row=row, column=6, value=vat_total)
            sheet.cell(row=row, column=7, value=invoice_number)

            workbook.save(self.location)
            QMessageBox.information(
                self, 'Information', 'Invoice saved successfully!')

            # تحديث البيانات في قاعدة البيانات بعد الحفظ
            insert_extracted_data(self.db_connection, self.current_detection_id, self.image_path,
                                  vendor_name, date, vat_id, invoice_total, vat_total, invoice_number)
        except Exception as e:
            QMessageBox.critical(
                self, 'Error', f'Failed to save invoice: {str(e)}')


if __name__ == '__main__':
    app = QApplication([])
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
